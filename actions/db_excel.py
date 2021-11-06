import openpyxl
import os.path as path

file_name_answer = 'E:\ChatBotRasa\ChatBot\ChatBot_v1\data_local\data_bot.xlsx'
file_name_question = 'E:\ChatBotRasa\ChatBot\ChatBot_v1\data_local\customer_question.xlsx'


# Tạo file excel mới
def create_excel_file(input_detail, output_excel_path):
    # Xác định số hàng và cột lớn nhất trong file excel cần tạo
    row = len(input_detail)
    column = len(input_detail[0])

    # Tạo một workbook mới và active nó
    wb_ = openpyxl.Workbook()
    ws = wb_.active

    # Dùng vòng lặp for để ghi nội dung từ input_detail vào file Excel
    for i in range(0, row):
        for j in range(0, column):
            v = input_detail[i][j]
            ws.cell(column=j + 1, row=i + 1, value=v)

    # Lưu lại file Excel
    wb_.save(output_excel_path)
    wb_.close()


# copy các work sheet sang file mới
def copy_work_sheet_not_data(output_excel_path, input_excel_path):
    # get work_sheet on file input
    wb = openpyxl.load_workbook(input_excel_path)
    g_sheet = wb.sheetnames
    # print(g_sheet)
    # Tạo một workbook mới và active nó
    wb_ = openpyxl.Workbook()
    for sheet in range(len(g_sheet)):
        ws = wb_.create_sheet(g_sheet[sheet])
        input_detail = read_questions(wb[g_sheet[sheet]])
        # Xác định số hàng và cột lớn nhất trong file excel cần tạo
        row = len(input_detail)
        column = len(input_detail[0])
        # Dùng vòng lặp for để ghi nội dung từ input_detail vào file Excel
        for i in range(0, row):
            for j in range(0, len(input_detail[i])):
                v = input_detail[i][j]
                ws.cell(column=j + 1, row=i + 1, value=v)
    wb_.remove(wb_['Sheet'])
    # Lưu lại file Excel
    wb_.save(output_excel_path)
    wb_.close()
    wb.close()


# ghi các message khách hàng vào file
def write_excel_file(intent_name, message):
    if path.isfile(file_name_question) is False:
        copy_work_sheet_not_data(file_name_question, file_name_answer)
    try:
        wb = openpyxl.load_workbook(file_name_question)
        g_sheet = wb.sheetnames
        tq = read_topic_question(intent_name)
        sheet = wb[g_sheet[tq.get('topic') - 1]]
        max_row = sheet.max_row + 2
        col = tq.get('question') + 1
        intent_ = sheet.cell(row=1, column=col).value
        if intent_ is None:
            # intent chưa có trong file thì tạo intent đó
            sheet.cell(row=1, column=col, value=intent_name)
            sheet.cell(row=2, column=col, value=message)
        else:
            # Tìm hàng trống điền message của khách
            for row in range(1, max_row):
                if sheet.cell(row=row, column=col).value is None:
                    sheet.cell(row=row, column=col, value=message)
                    break
        wb.save(file_name_question)
        wb.close()
    except Exception as e:
        print('Error', str(e))


# lấy list work sheet từ file
def read_excel_file(file_name=file_name_answer):
    try:
        wb = openpyxl.load_workbook(file_name)
        g_sheet = wb.sheetnames
        list_sheet = []
        for i in g_sheet:
            list_sheet.append(wb[i])
        wb.close()
        return list_sheet
    except Exception as e:
        print('Error read_excel_file:', str(e))
        return False


# lấy câu trả lời từ work sheet tương ứng với intent_name
def read_work_sheet(list_work_sheet, intent_name):
    tq = read_topic_question(intent_name)
    work_sheet = list_work_sheet[tq.get('topic') - 1]
    txt_value = work_sheet.cell(row=5 + tq.get('question'), column=5).value
    return txt_value


# trả về thứ tự topic và question tương ứng với intent_name
def read_topic_question(intent_name):
    text = str(intent_name).split('_')
    topic = text[0].strip('t')
    question = text[1].strip('q')
    return {'topic': int(topic), 'question': int(question)}


def read_questions(work_sheet):
    data = [['ID CÂU HỎI'], [1], [2], [3]]
    max_row = work_sheet.max_row
    for j in range(6, max_row + 1):
        intent = work_sheet.cell(row=j, column=3).value
        question = work_sheet.cell(row=j, column=4).value
        if question is not None:
            data[0].append(intent)
            data[1].append(question)
    return data


if __name__ == "__main__":
    var = read_excel_file()
    # print(read_questions(var[0]))
    print(var)
    # copy_work_sheet_not_data(file_name_question, file_name_answer)
# txt = read_work_sheet(var, "t1_q1")
# print(txt)
# #
# write_excel_file('t2_q4', "ok")
# write_excel_file('t2_q3', "ok")
# data = [['ID CÂU HỎI'], [1], [2], [3]]
# copy_work_sheet_not_data(data, file_name_question, file_name_answer)
